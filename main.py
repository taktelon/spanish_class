import random
import openpyxl as xl


class UniqueCounter:
    def __init__(self, min_num, max_num):
        self.min_num = min_num
        self.max_num = max_num
        self.unique_counter = set()

    def get_unique_number(self):
        for i in range(10):
            selected_number = random.randint(self.min_num, self.max_num)
            self.unique_counter.add(selected_number)
            if selected_number not in self.unique_counter:
                self.unique_counter.add(selected_number)
                break
            else:
                selected_number = random.randint(self.min_num, self.max_num)
        return selected_number


reflection = {
    "yo": "me",
    "vos": "te",
    "el/ella/usted": "se",
    "nosotros/nosotras": "nos",
    "ellos/ellas/ustedes": "se"
}

suffix_ar = {
    "yo": "o",
    "vos": "ás",
    "el/ella/usted": "a",
    "nosotros/nosotras": "amos",
    "ellos/ellas/ustedes": "an"
}
suffix_er = {
    "yo": "o",
    "vos": "és",
    "el/ella/usted": "e",
    "nosotros/nosotras": "emos",
    "ellos/ellas/ustedes": "en"
}
suffix_ir = {
    "yo": "o",
    "vos": "ís",
    "el/ella/usted": "e",
    "nosotros/nosotras": "imos",
    "ellos/ellas/ustedes": "en"
}

# initialize vers data sheel
wb = xl.load_workbook("Spanish.xlsx")
sheet = wb["verbs"]
sheet_words = wb["words"]
cell_yo = sheet.cell(1, 2)
max_row = sheet.max_row
max_row_words = sheet_words.max_row
max_column = sheet.max_column - 1
print(f"total {max_row - 1} verbs to practise")


def get_reflect_verb(verb_reflect, sub):
    verb_reflect_formed = get_regular_verb_form(verb_reflect, sub)
    return f"{reflection[sub]} {verb_reflect_formed}"


def get_regular_verb_form(verb_regular, sub):
    suf = verb_regular[-2:]
    head = verb_regular[:-2]
    if suf == "ar":
        head += suffix_ar[sub]
    elif suf == "er":
        head += suffix_er[sub]
    elif suf == "ir":
        head += suffix_ir[sub]
    elif suf == "se":
        head = get_reflect_verb(head, sub)
    else:
        head = "unknown"
    return head


def quiz_creator_verbs(quiz_row):
    quiz_column = random.randint(2, max_column)
    quiz_subject = sheet.cell(1, quiz_column).value
    quiz_verb = sheet.cell(quiz_row, 1).value
    quiz_answer = sheet.cell(quiz_row, quiz_column).value
    if quiz_answer is None:
        quiz_answer = get_regular_verb_form(quiz_verb, quiz_subject)
    return quiz_subject, quiz_verb, quiz_answer


def quiz_creator_word(quiz_word_row):
    quiz_word_word = sheet_words.cell(quiz_word_row, 2).value
    quiz_word_answer = sheet_words.cell(quiz_word_row, 1).value
    return "WORD", quiz_word_word, quiz_word_answer


your_choice = input("Chose [v]erbs or [w]ords: ").lower()
if your_choice == "v":
    choice = "verbs"
    counter = UniqueCounter(2, max_row)
elif your_choice == "w":
    choice = "words"
    counter = UniqueCounter(1, max_row_words)
else:
    print("Sorry, you don't feel learning...")
    exit(1)

while True:
    # subject, verb, answer = quiz_creator()
    if choice == "verbs":
        subject, verb, answer = quiz_creator_verbs(counter.get_unique_number())
    elif choice == "words":
        subject, verb, answer = quiz_creator_word(counter.get_unique_number())
    your_answer = input(f"[{subject}][{verb}] -> ").lower()
    if your_answer == "quit":
        break
    elif your_answer == answer:
        print("Correct!")
    else:
        print(f"Wrong! Should be {answer}!")
